import tkinter as tk
from tkinter import ttk
from tkinter import StringVar
from tkinter import END
from modulos.CRUD import *
from modulos.import_export import cal2fecha
from datetime import datetime
from tkcalendar import Calendar, DateEntry
import os
from openpyxl import Workbook
#exportar a excel https://es.stackoverflow.com/questions/36060/pasar-los-datos-de-una-lista-a-una-hoja-de-c%C3%A1lculo-mediante-dos-celdas-de-la-ho
class PaginaResumen(tk.Frame):
    def __init__(self, padre,  controlador):

        #####################################
        #CARGA INICIAL DEL OBJETO PÁGINA#
        #####################################
        tk.Frame.__init__(self, padre)
        ##setup para archivos excel
        self.controlador =  controlador
        self.ruta=controlador.ruta
        self.archivo="Asesor.xlsx"
        self.salida = os.path.join(self.ruta,self.archivo)
        #####################################
        #GRÁFICOS#
        #####################################
        label = tk.Label(self, text="Resumen de Facturas", font=controlador.fuente_titulo)
        label.pack(side="top", fill="x", pady=10)
        button = tk.Button(self, text="Ir al principio",
                           command=lambda:  controlador.mostrar_marco("PaginaInicial"))
        self.etiqueta_fecha_inicio = tk.Label(self, text="Fecha inicial para el EXCEL")
        self.cal_inicio = DateEntry(self, width=12, background='darkblue',
                        foreground='white', borderwidth=2, locale='es_ES')
        self.cal_inicio.set_date(datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),1,1))

        self.etiqueta_fecha_final = tk.Label(self, text="Fecha final para el EXCEL")
        self.cal_fin = DateEntry(self, width=12, background='darkblue',
                        foreground='white', borderwidth=2, locale='es_ES')
        self.cal_fin.set_date(datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),12,31))
     
        button2 = tk.Button(self, text="Facturado entre fechas",
                           command=lambda: self.trimestral(self.cal_inicio.get(),self.cal_fin.get()))
        button3 = tk.Button(self, text="Facturado anual",
                           command=lambda: self.trimestral(datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),1,1),datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),12,31)))
        
        button4 = tk.Button(self, text="Facturado 1T",
                           command=lambda: self.trimestral(datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),1,1),datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),3,31)))
        button5 = tk.Button(self, text="Facturado 2T",
                           command=lambda: self.trimestral(datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),4,1),datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),6,30)))
        button6 = tk.Button(self, text="Facturado 3T",
                           command=lambda: self.trimestral(datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),6,30),datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),9,30)))
        button7 = tk.Button(self, text="Facturado 4T",
                           command=lambda: self.trimestral(datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),10,1),datetime(2000+int(self.controlador.marcos["PaginaInicial"].configuracion["año"]),12,31)))
        
        button.pack()
        self.etiqueta_fecha_inicio.pack()
        self.cal_inicio.pack()
        
        self.etiqueta_fecha_final.pack()
        self.cal_fin.pack()

        button2.pack()
        button3.pack()
        button4.pack()
        button5.pack()
        button6.pack()
        button7.pack()
        datos_cliente_inicial=[('','Clientes contado','','','','','','','')]


        try:
            crearCliente(datos_cliente_inicial)
        except:
            print('Cliente ya creado')
        try:
            self.controlador.marcos['PaginaInicial'].actualizar()#hasta que no están todas las páginas creadas no puedo asignar la numeración automática, esta es la última página en inicializarse, por eso lo pongo aquí
        except:
            print('Sin id_factura')

        #####################################
        #FUNCIONES#
        #####################################
    def trimestral(self, fecha_inicial, fecha_final):
        #facturas = leerTodo("FACTURA")
        #clientes = leerTodo("CLIENTE")

        consulta = todoFacturasClientes("CODIGO_FACTURA", "FECHA_FACTURA", fecha_inicial, fecha_final)
        
        print ("Esta es la consulta que recibe antes de pasar a excel: \n",consulta)
        excel = Workbook()
        hoja= excel.active
        hoja.title = "Facturado"
       

        fila_ptr = 1 #los utilizo a modo de punteros
        columna_ptr = 1
        #1ª fila
        hoja.cell(column=columna_ptr, row=fila_ptr, value=f"Declarado entre {fecha_inicial} y {fecha_final}")
        fila_ptr +=1
        #2ª fila
        hoja.cell(column=columna_ptr, row=fila_ptr, value="Factura")
        columna_ptr +=1
        hoja.cell(column=columna_ptr, row=fila_ptr, value="Fecha")
        columna_ptr +=1
        hoja.cell(column=columna_ptr, row=fila_ptr, value="Cliente")
        columna_ptr +=1
        hoja.cell(column=columna_ptr, row=fila_ptr, value="DNI")
        columna_ptr +=1
        hoja.cell(column=columna_ptr, row=fila_ptr, value="Fecha Factura")
        columna_ptr +=1
        hoja.cell(column=columna_ptr, row=fila_ptr, value="Servicios realizados")
        columna_ptr +=1
        hoja.cell(column=columna_ptr, row=fila_ptr, value="€")
        columna_ptr +=1
        hoja.cell(column=columna_ptr, row=fila_ptr, value="IVA %")
        columna_ptr +=1
        hoja.cell(column=columna_ptr, row=fila_ptr, value="Descuento % ")
        columna_ptr +=1
        hoja.cell(column=columna_ptr, row=fila_ptr, value="Total facturado €(Impuestos y descuentos incluidos")
        columna_ptr =1
        fila_ptr +=1
        #3ª fila
        factura_leida = "ninguna"
        
        acumulado = 0
        servicios_contados=1
        for info in consulta:
            if factura_leida == info[1]:

                servicios_contados+=1
                acumulado +=info[-2]

                print ("se tiene en cuenta ", acumulado)
                hoja.cell(column=columna_ptr+5, row=fila_ptr-1, value=f"{servicios_contados} Tratamientos")
                hoja.cell(column=columna_ptr+6, row=fila_ptr-1, value=acumulado)
                hoja.cell(column=columna_ptr+9, row=fila_ptr-1, value=round((acumulado*(1+info[3]/100))*(1-info[4]/100),2))
                continue
            else:
                acumulado = 0
                servicios_contados=1
                

            
            factura_leida = info[1]
            hoja.cell(column=columna_ptr, row=fila_ptr, value=factura_leida)#codigo
            columna_ptr +=1
            hoja.cell(column=columna_ptr, row=fila_ptr, value=info[2])#fecha
            columna_ptr +=1
            hoja.cell(column=columna_ptr, row=fila_ptr, value=info[9]+' '+info[10]) #nombre
            columna_ptr +=1
            hoja.cell(column=columna_ptr, row=fila_ptr, value=info[8])#DNI
            columna_ptr +=1
            hoja.cell(column=columna_ptr, row=fila_ptr, value=info[-4])#fecha último servicio
            columna_ptr +=1
            hoja.cell(column=columna_ptr, row=fila_ptr, value=info[-3])#tratamiento
            columna_ptr +=1
            hoja.cell(column=columna_ptr, row=fila_ptr, value=info[-2])#Precio del tratamiento
            columna_ptr +=1
            hoja.cell(column=columna_ptr, row=fila_ptr, value=info[3])#IVA
            columna_ptr +=1
            hoja.cell(column=columna_ptr, row=fila_ptr, value=info[4])#Descuento
            columna_ptr +=1
            acumulado += info[-2]*(1+float(info[3])/100)*(1-float(info[4])/100)
            acumulado = round(acumulado, 2)
            hoja.cell(column=columna_ptr, row=fila_ptr, value=acumulado)
            columna_ptr =1
            fila_ptr +=1
            
            print(info)

        #Grabar el archivo
        excel.save(filename=self.salida)
        comando = "start excel.exe "+"\""+ self.salida +"\""+" &"
        os.system(comando)